"""
services/jnto_fetcher.py — JNTO 월별 방일 외래관광객 (한국) 자동 수집

JNTO 통계 페이지에서 최신 Excel URL을 스크래핑 후 다운로드,
韓国 행 데이터를 파싱하여 tourism_monthly 테이블에 저장.

대상: 한국인 방일 관광객 (訪日外客数 韓国)
빈도: 월 1회 자동 실행 (app.py 스케줄러)
ENV: 없음 (공개 데이터)
"""
import re
import io
import urllib.request
import urllib.error

from db import get_db

_STATS_PAGE = "https://www.jnto.go.jp/statistics/data/visitors-statistics/"
_FALLBACK_URL = "https://www.jnto.go.jp/statistics/data/visitors-statistics/"

MONTH_COLS = [2, 4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24]


def _find_xlsx_url() -> str | None:
    """통계 페이지 HTML에서 최신 .xlsx 링크를 찾아 반환."""
    try:
        req = urllib.request.Request(_STATS_PAGE, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=20) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        print(f"[JNTO] 페이지 로드 실패: {e}", flush=True)
        return None

    # /_files/YYYYMMDD_NNNN-N.xlsx 패턴
    matches = re.findall(r'/_files/\d{8}_\d{4}-\d\.xlsx', html)
    if not matches:
        matches = re.findall(r'/statistics/data/_files/[^"\']+\.xlsx', html)
    if not matches:
        return None

    path = matches[0]
    if path.startswith("/_files/"):
        return "https://www.jnto.go.jp/statistics/data" + path
    return "https://www.jnto.go.jp" + path


def _download_xlsx(url: str) -> bytes | None:
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.read()
    except Exception as e:
        print(f"[JNTO] Excel 다운로드 실패 ({url}): {e}", flush=True)
        return None


def _parse_and_save(xlsx_bytes: bytes) -> int:
    try:
        import openpyxl
    except ImportError:
        print("[JNTO] openpyxl 미설치 — pip install openpyxl", flush=True)
        return 0

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    conn = get_db()
    saved = 0

    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 8:
            continue

        # 韓国 행 찾기 (보통 index 6 = 7번째 행)
        korea_row = None
        for row in rows[4:15]:
            first = str(row[0] or "").strip()
            if "韓国" in first or "한국" in first.lower():
                korea_row = row
                break

        if korea_row is None:
            continue

        for month_0, col_idx in enumerate(MONTH_COLS, start=1):
            if col_idx >= len(korea_row):
                break
            val = korea_row[col_idx]
            try:
                visitors = int(val)
            except (TypeError, ValueError):
                continue
            if visitors <= 0:
                continue

            ym = f"{year}-{month_0:02d}"
            conn.execute(
                """INSERT INTO tourism_monthly (year_month, country, visitors, source, fetched_at)
                   VALUES (?, 'japan', ?, 'jnto', datetime('now','localtime'))
                   ON CONFLICT(year_month, country) DO UPDATE
                   SET visitors=excluded.visitors, source=excluded.source,
                       fetched_at=excluded.fetched_at""",
                (ym, visitors)
            )
            saved += 1

    conn.commit()
    conn.close()
    return saved


def fetch_jnto() -> int:
    """JNTO 최신 Excel을 가져와 DB 저장. 저장 건수 반환."""
    url = _find_xlsx_url()
    if not url:
        print("[JNTO] xlsx URL 탐지 실패", flush=True)
        return 0

    print(f"[JNTO] 다운로드: {url}", flush=True)
    xlsx_bytes = _download_xlsx(url)
    if not xlsx_bytes:
        return 0

    saved = _parse_and_save(xlsx_bytes)
    print(f"[JNTO] {saved}개월분 저장", flush=True)
    return saved


def fetch_jnto_from_url(url: str) -> int:
    """지정 URL에서 직접 Excel 다운로드 후 저장."""
    print(f"[JNTO] 직접 다운로드: {url}", flush=True)
    xlsx_bytes = _download_xlsx(url)
    if not xlsx_bytes:
        return 0
    saved = _parse_and_save(xlsx_bytes)
    print(f"[JNTO] {saved}개월분 저장", flush=True)
    return saved
