"""
routes/monitor.py — 대시보드, 게시글 상세, 기본 API 라우트
"""
import os
import threading
from datetime import datetime

from flask import Blueprint, jsonify, render_template, request

from config import KST
from db import get_db
from services.naver import collect_all
from services.analysis import process_unanalyzed
from services.email_svc import send_daily_report

monitor_bp = Blueprint("monitor", __name__)

# GLN 서비스 지원 국가 감지 (gln-content 의존성 없이 로컬 복사)
_COUNTRY_MAP = {
    # 국가명
    "태국": "thailand",    "방콕": "thailand",
    "일본": "japan",       "도쿄": "japan",       "오사카": "japan",
    "대만": "taiwan",      "타이베이": "taiwan",
    "베트남": "vietnam",   "호치민": "vietnam",    "하노이": "vietnam",
    "필리핀": "philippines", "마닐라": "philippines",
    "싱가포르": "singapore",
    "홍콩": "hongkong",
    "마카오": "macau",
    "중국": "china",       "베이징": "china",      "상하이": "china",
    "캄보디아": "cambodia", "프놈펜": "cambodia",
    "몽골": "mongolia",    "울란바토르": "mongolia",
    "라오스": "laos",
    "괌": "guam",
    "사이판": "saipan",
    # 베트남 도시
    "나트랑": "vietnam",   "냐짱": "vietnam",     "다낭": "vietnam",
    "달랏": "vietnam",     "다랏": "vietnam",     "푸꾸옥": "vietnam",
    "할롱": "vietnam",     "호이안": "vietnam",   "무이네": "vietnam",
    "빈펄": "vietnam",     "사파": "vietnam",     "붕따우": "vietnam",
    # 태국 도시
    "치앙마이": "thailand", "파타야": "thailand",  "푸켓": "thailand",
    "사무이": "thailand",   "끄라비": "thailand",  "후아힌": "thailand",
    # 일본 도시
    "교토": "japan",       "후쿠오카": "japan",    "삿포로": "japan",
    "오키나와": "japan",   "나고야": "japan",      "나하": "japan",
    "고베": "japan",       "요코하마": "japan",    "히로시마": "japan",
    # 대만 도시
    "가오슝": "taiwan",    "타이중": "taiwan",     "타이난": "taiwan",
    "화롄": "taiwan",
    # 필리핀 도시
    "세부": "philippines", "보라카이": "philippines", "다바오": "philippines",
    "팔라완": "philippines", "엘니도": "philippines",
    # 중국 도시
    "광저우": "china",     "선전": "china",        "청두": "china",
    "시안": "china",       "항저우": "china",      "구이린": "china",
    # 캄보디아 도시
    "씨엠립": "cambodia",  "앙코르": "cambodia",
    # 라오스 도시
    "비엔티안": "laos",    "루앙프라방": "laos",   "방비엥": "laos",
    # 괌
    "투몬": "guam",
}

COUNTRY_LABEL = {
    "thailand":    "태국",
    "japan":       "일본",
    "taiwan":      "대만",
    "vietnam":     "베트남",
    "philippines": "필리핀",
    "singapore":   "싱가포르",
    "hongkong":    "홍콩",
    "macau":       "마카오",
    "china":       "중국",
    "cambodia":    "캄보디아",
    "mongolia":    "몽골",
    "laos":        "라오스",
    "guam":        "괌사이판",
    "saipan":      "괌사이판",
}

COUNTRY_EMOJI = {
    "vietnam":     "🇻🇳", "china":       "🇨🇳", "hongkong":    "🇭🇰",
    "macau":       "🇲🇴", "philippines": "🇵🇭", "thailand":    "🇹🇭",
    "laos":        "🇱🇦", "japan":       "🇯🇵", "taiwan":      "🇹🇼",
    "mongolia":    "🇲🇳", "singapore":   "🇸🇬", "cambodia":    "🇰🇭",
    "guam":        "🏝️",  "saipan":      "🏝️",
}


def _detect_country(text: str) -> str:
    for kor, eng in _COUNTRY_MAP.items():
        if kor in text:
            return eng
    return ""


@monitor_bp.route("/")
def dashboard():
    conn = get_db()

    sentiment    = request.args.get("sentiment", "")
    category     = request.args.get("category", "")
    urgent       = request.args.get("urgent", "")
    country      = request.args.get("country", "")
    today_str    = datetime.now(KST).strftime("%Y-%m-%d")
    date_from    = request.args.get("date_from", today_str)
    date_to      = request.args.get("date_to", today_str)
    channel      = request.args.get("channel", "")
    reply_status = request.args.get("reply_status", "")

    query = """
        SELECT p.id, p.title, p.link, p.cafe_name, p.post_date, p.is_urgent,
               p.keyword, p.created_at, p.reply_status, p.status_updated_at,
               p.description,
               a.summary, a.category, a.sentiment, a.importance_score
        FROM posts p
        LEFT JOIN ai_analysis a ON p.id = a.post_id
        WHERE (a.is_relevant IS NULL OR a.is_relevant = 1)
    """
    args = []
    if sentiment:
        query += " AND a.sentiment = ?";      args.append(sentiment)
    if category:
        query += " AND a.category = ?";       args.append(category)
    if urgent == "1":
        query += " AND p.is_urgent = 1"
    if channel:
        query += " AND p.keyword LIKE ?";     args.append(f"{channel}/%")
    if reply_status:
        query += " AND p.reply_status = ?";   args.append(reply_status)
    if date_from:
        query += " AND DATE(p.created_at) >= ?"; args.append(date_from)
    if date_to:
        query += " AND DATE(p.created_at) <= ?"; args.append(date_to)
    if country:
        _targets = ('guam', 'saipan') if country == 'guam_saipan' else (country,)
        _kws = [k for k, v in _COUNTRY_MAP.items() if v in _targets]
        _cc  = " OR ".join(f"(p.title LIKE ? OR p.description LIKE ? OR p.cafe_name LIKE ?)" for _ in _kws)
        if _cc:
            query += f" AND ({_cc})"
            for _kw in _kws:
                args.extend([f"%{_kw}%", f"%{_kw}%", f"%{_kw}%"])

    page     = int(request.args.get("page", 1))
    per_page = 50
    offset   = (page - 1) * per_page

    count_query = """
        SELECT COUNT(*) FROM posts p
        LEFT JOIN ai_analysis a ON p.id = a.post_id
        WHERE (a.is_relevant IS NULL OR a.is_relevant = 1)
    """
    count_args = []
    if sentiment:
        count_query += " AND a.sentiment = ?";      count_args.append(sentiment)
    if category:
        count_query += " AND a.category = ?";       count_args.append(category)
    if urgent == "1":
        count_query += " AND p.is_urgent = 1"
    if channel:
        count_query += " AND p.keyword LIKE ?";     count_args.append(f"{channel}/%")
    if reply_status:
        count_query += " AND p.reply_status = ?";   count_args.append(reply_status)
    if date_from:
        count_query += " AND DATE(p.created_at) >= ?"; count_args.append(date_from)
    if date_to:
        count_query += " AND DATE(p.created_at) <= ?"; count_args.append(date_to)
    if country and _kws and _cc:
        count_query += f" AND ({_cc})"
        for _kw in _kws:
            count_args.extend([f"%{_kw}%", f"%{_kw}%", f"%{_kw}%"])

    total       = conn.execute(count_query, count_args).fetchone()[0]
    total_pages = max(1, (total + per_page - 1) // per_page)

    query += f" ORDER BY p.created_at DESC LIMIT {per_page} OFFSET {offset}"
    _rows = conn.execute(query, args).fetchall()
    posts = []
    for r in _rows:
        d = dict(r)
        d["country"] = _detect_country(
            (d.get("title") or "") + " " + (d.get("description") or "") + " " + (d.get("cafe_name") or "")
        )
        posts.append(d)

    stats_where = "WHERE 1=1"
    stats_args  = []
    if date_from:
        stats_where += " AND DATE(created_at) >= ?"; stats_args.append(date_from)
    if date_to:
        stats_where += " AND DATE(created_at) <= ?"; stats_args.append(date_to)

    channel_counts = {}
    for ch in ["카페", "블로그", "뉴스"]:
        cnt = conn.execute(
            f"SELECT COUNT(*) FROM posts {stats_where} AND keyword LIKE ?",
            stats_args + [f"{ch}/%"]
        ).fetchone()[0]
        channel_counts[ch] = cnt

    stats = {
        "today":    conn.execute(f"SELECT COUNT(*) FROM posts {stats_where}", stats_args).fetchone()[0],
        "urgent":   conn.execute(f"SELECT COUNT(*) FROM posts {stats_where} AND is_urgent=1", stats_args).fetchone()[0],
        "negative": conn.execute("SELECT COUNT(*) FROM ai_analysis WHERE sentiment='negative'").fetchone()[0],
        "total":    conn.execute("SELECT COUNT(*) FROM posts").fetchone()[0],
    }
    conn.close()
    report_to = os.getenv("REPORT_TO", "")
    return render_template(
        "dashboard.html", posts=posts, stats=stats,
        channel_counts=channel_counts,
        report_to=report_to,
        today_str=today_str,
        country_label=COUNTRY_LABEL,
        country_emoji=COUNTRY_EMOJI,
        filters={"sentiment": sentiment, "category": category,
                 "urgent": urgent, "date_from": date_from, "date_to": date_to,
                 "channel": channel, "reply_status": reply_status, "country": country},
        page=page, total_pages=total_pages, total=total
    )


@monitor_bp.route("/post/<int:post_id>")
def post_detail(post_id):
    conn = get_db()
    post = conn.execute(
        """SELECT p.*, a.summary, a.category, a.sentiment, a.importance_score
           FROM posts p LEFT JOIN ai_analysis a ON p.id=a.post_id
           WHERE p.id=?""", (post_id,)
    ).fetchone()
    replies = conn.execute(
        "SELECT type, content FROM draft_replies WHERE post_id=?", (post_id,)
    ).fetchall()
    conn.close()
    return render_template("post_detail.html", post=post, replies=replies)


@monitor_bp.route("/insights")
def insights():
    return render_template("insights.html")


@monitor_bp.route("/api/status/<int:post_id>", methods=["POST"])
def api_update_status(post_id):
    data   = request.get_json(silent=True) or {}
    status = data.get("status", "")
    if status not in ["미확인", "확인완료", "답변완료"]:
        return jsonify({"error": "잘못된 상태값"}), 400
    now  = datetime.now(KST).strftime("%Y-%m-%d %H:%M")
    conn = get_db()
    conn.execute(
        "UPDATE posts SET reply_status=?, status_updated_at=? WHERE id=?",
        (status, now, post_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"status": status, "updated_at": now})


@monitor_bp.route("/api/status/bulk", methods=["POST"])
def api_bulk_update_status():
    data   = request.get_json(silent=True) or {}
    ids    = data.get("ids", [])
    status = data.get("status", "")
    if status not in ["미확인", "확인완료", "답변완료"]:
        return jsonify({"error": "잘못된 상태값"}), 400
    if not ids or not isinstance(ids, list):
        return jsonify({"error": "ids 필요"}), 400
    now  = datetime.now(KST).strftime("%Y-%m-%d %H:%M")
    conn = get_db()
    placeholders = ",".join("?" * len(ids))
    conn.execute(
        f"UPDATE posts SET reply_status=?, status_updated_at=? WHERE id IN ({placeholders})",
        [status, now] + ids
    )
    conn.commit()
    conn.close()
    return jsonify({"updated": len(ids), "status": status})


@monitor_bp.route("/api/collect", methods=["POST"])
def api_collect():
    threading.Thread(target=collect_all, daemon=True).start()
    return jsonify({"status": "수집 시작됨"})


@monitor_bp.route("/api/process", methods=["POST"])
def api_process():
    threading.Thread(target=process_unanalyzed, daemon=True).start()
    return jsonify({"status": "AI 분석 시작됨"})


@monitor_bp.route("/api/report", methods=["POST"])
def api_report():
    if os.getenv("DISABLE_EMAIL_SEND", "false").lower() == "true":
        return jsonify({"status": "이메일 발송 비활성화됨 (DISABLE_EMAIL_SEND=true)"})
    import datetime as _dt
    data    = request.get_json(silent=True) or {}
    weekday = _dt.datetime.now().weekday()
    setting_key = "report_to_weekend" if weekday >= 5 else "report_to_weekday"
    to = data.get("to", "").strip() or (get_setting(setting_key) or os.getenv("REPORT_TO", ""))
    print(f"[API] 리포트 발송 요청 — 수신자: {to}", flush=True)
    threading.Thread(target=send_daily_report, args=(to,), daemon=True).start()
    return jsonify({"status": "리포트 발송 시작됨"})


def _insights_date_clause(date_from, date_to, alias=""):
    """date_from/date_to 기반 WHERE 절 조각과 파라미터 반환."""
    p = f"{alias}." if alias else ""
    clauses, args = [], []
    if date_from:
        clauses.append(f"DATE({p}created_at) >= ?"); args.append(date_from)
    if date_to:
        clauses.append(f"DATE({p}created_at) <= ?"); args.append(date_to)
    return (" AND " + " AND ".join(clauses)) if clauses else "", args


@monitor_bp.route("/api/insights")
def api_insights():
    conn      = get_db()
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    # fallback: days 파라미터 (하위 호환)
    if not date_from and not date_to:
        days      = int(request.args.get("days", 7))
        date_from = datetime.now(KST).strftime("%Y-%m-%d") if days == 1 else ""
        w_clause  = f"created_at >= DATE('now', '-{days} days', 'localtime')"
        w_args    = []
    else:
        w_clause, w_args = "", []
        if date_from:
            w_clause += ("AND " if w_clause else "") + "DATE(created_at) >= ?"
            w_args.append(date_from)
        if date_to:
            w_clause += (" AND " if w_clause else "") + "DATE(created_at) <= ?"
            w_args.append(date_to)

    where = f"WHERE {w_clause}" if w_clause else "WHERE 1=1"

    daily = conn.execute(f"""
        SELECT DATE(created_at) as day,
               SUM(CASE WHEN keyword LIKE '카페/%' THEN 1 ELSE 0 END) as cafe,
               SUM(CASE WHEN keyword LIKE '블로그/%' THEN 1 ELSE 0 END) as blog,
               SUM(CASE WHEN keyword LIKE '뉴스/%' THEN 1 ELSE 0 END) as news,
               COUNT(*) as total
        FROM posts {where}
        GROUP BY DATE(created_at)
        ORDER BY day ASC
    """, w_args).fetchall()

    sentiment = conn.execute(f"""
        SELECT a.sentiment, COUNT(*) as cnt
        FROM ai_analysis a
        JOIN posts p ON a.post_id = p.id
        WHERE a.sentiment IS NOT NULL
          {'AND ' + w_clause.replace('created_at', 'p.created_at') if w_clause else ''}
        GROUP BY a.sentiment
    """, w_args).fetchall()

    keywords = conn.execute(f"""
        SELECT SUBSTR(keyword, INSTR(keyword, '/') + 1) as kw,
               COUNT(*) as cnt
        FROM posts {where}
          AND keyword IS NOT NULL
        GROUP BY kw ORDER BY cnt DESC LIMIT 10
    """, w_args).fetchall()

    total    = conn.execute(f"SELECT COUNT(*) FROM posts {where}", w_args).fetchone()[0]
    urgent   = conn.execute(f"SELECT COUNT(*) FROM posts {where} AND is_urgent=1", w_args).fetchone()[0]
    negative = conn.execute(f"""
        SELECT COUNT(*) FROM ai_analysis a JOIN posts p ON a.post_id=p.id
        WHERE a.sentiment='negative'
          {'AND ' + w_clause.replace('created_at', 'p.created_at') if w_clause else ''}
    """, w_args).fetchone()[0]
    conn.close()

    return jsonify({
        "daily":     [dict(r) for r in daily],
        "sentiment": [dict(r) for r in sentiment],
        "keywords":  [dict(r) for r in keywords],
        "summary":   {"total": total, "urgent": urgent, "negative": negative}
    })


@monitor_bp.route("/api/stats")
def api_stats():
    conn  = get_db()
    stats = {
        "today":    conn.execute("SELECT COUNT(*) FROM posts WHERE DATE(created_at)=DATE('now','localtime')").fetchone()[0],
        "urgent":   conn.execute("SELECT COUNT(*) FROM posts WHERE is_urgent=1 AND DATE(created_at)=DATE('now','localtime')").fetchone()[0],
        "negative": conn.execute("SELECT COUNT(*) FROM ai_analysis WHERE sentiment='negative'").fetchone()[0],
        "total":    conn.execute("SELECT COUNT(*) FROM posts").fetchone()[0],
    }
    conn.close()
    return jsonify(stats)


@monitor_bp.route("/api/insights/issues")
def api_insights_issues():
    """기간 내 주목 이슈 목록 — 인사이트 이슈 섹션용."""
    conn      = get_db()
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    filt      = request.args.get("filter", "all")  # all | urgent | negative | no_content

    w_parts, w_args = ["1=1"], []
    if date_from:
        w_parts.append("DATE(p.created_at) >= ?"); w_args.append(date_from)
    if date_to:
        w_parts.append("DATE(p.created_at) <= ?"); w_args.append(date_to)
    if filt == "urgent":
        w_parts.append("p.is_urgent = 1")
    elif filt == "negative":
        w_parts.append("a.sentiment = 'negative'")

    where = " AND ".join(w_parts)

    rows = conn.execute(f"""
        SELECT p.id, p.title, p.link, p.cafe_name, p.keyword, p.created_at,
               p.is_urgent, p.reply_status,
               a.summary, a.sentiment, a.importance_score, a.category,
               (SELECT COUNT(*) FROM content_drafts cd WHERE cd.source_post_id = p.id) as content_count
        FROM posts p
        LEFT JOIN ai_analysis a ON p.id = a.post_id
        WHERE {where}
        ORDER BY a.importance_score DESC, p.created_at DESC
        LIMIT 50
    """, w_args).fetchall()

    if filt == "no_content":
        rows = [r for r in rows if r["content_count"] == 0]

    conn.close()
    return jsonify([dict(r) for r in rows])


@monitor_bp.route("/api/insights/advanced")
def api_insights_advanced():
    conn      = get_db()
    date_from = request.args.get("date_from", "")
    date_to   = request.args.get("date_to", "")
    # fallback: days 파라미터 (하위 호환)
    if not date_from and not date_to:
        days = int(request.args.get("days", 7))
    else:
        days = None

    # ── 날짜 범위 WHERE 조각 생성 ─────────────────────────────────────────────
    from datetime import timedelta
    if days is not None:
        today     = datetime.now(KST).date()
        date_to   = today.strftime("%Y-%m-%d")
        date_from = (today - timedelta(days=days - 1)).strftime("%Y-%m-%d")
        prev_to   = (today - timedelta(days=days)).strftime("%Y-%m-%d")
        prev_from = (today - timedelta(days=days * 2 - 1)).strftime("%Y-%m-%d")
    else:
        from datetime import date as _date
        d_from = datetime.strptime(date_from, "%Y-%m-%d").date() if date_from else _date(2000,1,1)
        d_to   = datetime.strptime(date_to,   "%Y-%m-%d").date() if date_to   else datetime.now(KST).date()
        span   = (d_to - d_from).days + 1
        prev_to   = (d_from - timedelta(days=1)).strftime("%Y-%m-%d")
        prev_from = (d_from - timedelta(days=span)).strftime("%Y-%m-%d")

    def _w(alias="p"):
        parts = []
        if date_from: parts.append(f"DATE({alias}.created_at) >= '{date_from}'")
        if date_to:   parts.append(f"DATE({alias}.created_at) <= '{date_to}'")
        return (" AND " + " AND ".join(parts)) if parts else ""

    def _pw(alias="p"):
        return f" AND DATE({alias}.created_at) >= '{prev_from}' AND DATE({alias}.created_at) <= '{prev_to}'"

    # ── 이번 기간 vs 이전 기간 비교 ───────────────────────────────────────────
    total_this = conn.execute(f"SELECT COUNT(*) FROM posts p WHERE 1=1{_w('p')}").fetchone()[0]
    total_prev = conn.execute(f"SELECT COUNT(*) FROM posts p WHERE 1=1{_pw('p')}").fetchone()[0]
    urgent_this = conn.execute(f"SELECT COUNT(*) FROM posts p WHERE p.is_urgent=1{_w('p')}").fetchone()[0]
    urgent_prev = conn.execute(f"SELECT COUNT(*) FROM posts p WHERE p.is_urgent=1{_pw('p')}").fetchone()[0]
    neg_this = conn.execute(f"""
        SELECT COUNT(*) FROM ai_analysis a JOIN posts p ON a.post_id=p.id
        WHERE a.sentiment='negative'{_w('p')}
    """).fetchone()[0]
    neg_prev = conn.execute(f"""
        SELECT COUNT(*) FROM ai_analysis a JOIN posts p ON a.post_id=p.id
        WHERE a.sentiment='negative'{_pw('p')}
    """).fetchone()[0]

    # ── 브랜드 헬스 스코어 ────────────────────────────────────────────────────
    sent_rows = conn.execute(f"""
        SELECT a.sentiment, COUNT(*) as cnt
        FROM ai_analysis a JOIN posts p ON a.post_id=p.id
        WHERE a.sentiment IS NOT NULL{_w('p')}
        GROUP BY a.sentiment
    """).fetchall()
    sentiment_map = {r["sentiment"]: r["cnt"] for r in sent_rows}
    s_total      = sum(sentiment_map.values()) or 1
    pos_pct      = round(sentiment_map.get("positive", 0) / s_total * 100)
    neg_pct      = round(sentiment_map.get("negative", 0) / s_total * 100)
    health_score = round(pos_pct * 0.6 + (100 - neg_pct) * 0.4)

    # ── 키워드별 감성 트렌드 ─────────────────────────────────────────────────
    kw_sentiment = conn.execute(f"""
        SELECT SUBSTR(p.keyword, INSTR(p.keyword, '/') + 1) as kw,
               a.sentiment, COUNT(*) as cnt
        FROM posts p
        LEFT JOIN ai_analysis a ON p.id = a.post_id
        WHERE p.keyword IS NOT NULL AND a.sentiment IS NOT NULL{_w('p')}
        GROUP BY kw, a.sentiment ORDER BY kw, a.sentiment
    """).fetchall()

    kw_sent_map = {}
    for r in kw_sentiment:
        kw = r["kw"]
        if kw not in kw_sent_map:
            kw_sent_map[kw] = {"positive": 0, "neutral": 0, "negative": 0, "total": 0}
        kw_sent_map[kw][r["sentiment"]] = r["cnt"]
        kw_sent_map[kw]["total"] += r["cnt"]
    kw_sentiment_list = [
        {"kw": kw, **vals}
        for kw, vals in sorted(kw_sent_map.items(), key=lambda x: -x[1]["total"])
    ]

    # ── 미처리 SLA 현황 ───────────────────────────────────────────────────────
    sla_stats = conn.execute("""
        SELECT
            COUNT(*) as total_unprocessed,
            SUM(CASE WHEN (julianday('now','localtime') - julianday(created_at)) * 24 >= 6  THEN 1 ELSE 0 END) as over_6h,
            SUM(CASE WHEN (julianday('now','localtime') - julianday(created_at)) * 24 >= 24 THEN 1 ELSE 0 END) as over_24h,
            ROUND(AVG((julianday('now','localtime') - julianday(created_at)) * 24), 1) as avg_hours
        FROM posts
        WHERE reply_status = '미확인'
    """).fetchone()

    # ── 응답률 ────────────────────────────────────────────────────────────────
    status_rows = conn.execute(f"""
        SELECT reply_status, COUNT(*) as cnt
        FROM posts p WHERE 1=1{_w('p')}
        GROUP BY reply_status
    """).fetchall()
    status_map = {r["reply_status"]: r["cnt"] for r in status_rows}
    total_s    = sum(status_map.values()) or 1
    response_rate = round((status_map.get("확인완료", 0) + status_map.get("답변완료", 0)) / total_s * 100)

    conn.close()

    def pct_change(curr, prev):
        if prev == 0:
            return 100 if curr > 0 else 0
        return round(((curr - prev) / prev) * 100, 1)

    return jsonify({
        "comparison": {
            "total":  {"this": total_this,  "prev": total_prev,  "change": pct_change(total_this, total_prev)},
            "urgent": {"this": urgent_this, "prev": urgent_prev, "change": pct_change(urgent_this, urgent_prev)},
            "negative": {"this": neg_this,  "prev": neg_prev,    "change": pct_change(neg_this, neg_prev)},
        },
        "health_score": health_score,
        "kw_sentiment": kw_sentiment_list,
        "sla": {
            "total_unprocessed": sla_stats["total_unprocessed"] or 0,
            "over_6h":           sla_stats["over_6h"] or 0,
            "over_24h":          sla_stats["over_24h"] or 0,
            "avg_hours":         sla_stats["avg_hours"] or 0,
        },
        "response_rate": response_rate,
        "status_dist":   dict(status_map),
        "date_from": date_from,
        "date_to":   date_to,
    })


@monitor_bp.route("/api/report/weekly", methods=["POST"])
def api_weekly_report():
    if os.getenv("DISABLE_EMAIL_SEND", "false").lower() == "true":
        return jsonify({"status": "이메일 발송 비활성화됨 (DISABLE_EMAIL_SEND=true)"})
    from services.weekly_report import send_weekly_report
    from db import get_setting
    to = (get_setting("report_to_list") or os.getenv("REPORT_TO", "")).strip()
    threading.Thread(target=send_weekly_report, args=(to,), daemon=True).start()
    return jsonify({"status": "주간 리포트 발송 시작됨"})


@monitor_bp.route("/api/insights/tourism")
def api_insights_tourism():
    import os as _os
    from services.tourism_stats import fetch_recent_months, update_all
    has_key = bool(_os.getenv("KOSIS_API_KEY", "").strip())
    data = fetch_recent_months(13)
    if not data["months"]:
        if has_key:
            threading.Thread(target=update_all, daemon=True).start()
            return jsonify({"months": [], "countries": {}, "last_updated": "", "loading": True})
        return jsonify({"months": [], "countries": {}, "last_updated": "", "no_key": True})
    return jsonify(data)


# ── 월별 관광통계 API ─────────────────────────────────────────────

@monitor_bp.route("/api/insights/tourism-monthly")
def api_insights_tourism_monthly():
    n = int(request.args.get("n", 24))
    conn = get_db()
    rows = conn.execute(
        "SELECT year_month, country, visitors, fetched_at FROM tourism_monthly ORDER BY year_month"
    ).fetchall()
    conn.close()

    if not rows:
        return jsonify({"months": [], "countries": {}, "last_updated": ""})

    all_months = sorted({r["year_month"] for r in rows})
    months = all_months[-n:]
    year_idx = {m: i for i, m in enumerate(months)}

    countries: dict = {}
    for r in rows:
        ym = r["year_month"]
        if ym not in year_idx:
            continue
        code = r["country"]
        if code not in countries:
            countries[code] = [None] * len(months)
        countries[code][year_idx[ym]] = r["visitors"]

    last_updated = max((r["fetched_at"] or "") for r in rows)
    return jsonify({"months": months, "countries": countries, "last_updated": last_updated[:10]})


@monitor_bp.route("/api/admin/tourism-upload", methods=["POST"])
def api_tourism_upload():
    """Excel 파일 업로드로 tourism_monthly 벌크 저장.

    Form fields:
      file   — xlsx 파일 (openpyxl로 파싱, 형식: A열=YYYY-MM, B열=국가코드, C열=방문자수)
      country — 단일 국가 코드 (JNTO 형식 xlsx 처리 시에도 사용)
      mode   — 'jnto' | 'simple' (기본 'simple')
    """
    from flask import request as req

    mode = req.form.get("mode", "simple")
    country = req.form.get("country", "").strip()

    if "file" not in req.files:
        return jsonify({"ok": False, "error": "파일 없음"}), 400

    f = req.files["file"]
    xlsx_bytes = f.read()

    if mode == "jnto":
        from services.jnto_fetcher import _parse_and_save
        saved = _parse_and_save(xlsx_bytes)
        return jsonify({"ok": True, "saved": saved})

    if mode == "kto":
        # KTO DataLab Excel 형식:
        # 1행: 헤더 (A열=기간, B열~=국가명 in Korean)
        # 2행~: A열=기간(YYYY년 MM월 or 202401 etc.), B열~=방문자수
        _KTO_CMAP = {
            "베트남":"vietnam", "태국":"thailand", "일본":"japan", "대만":"taiwan",
            "필리핀":"philippines", "싱가포르":"singapore", "홍콩":"hongkong",
            "마카오":"macau", "중국":"china", "캄보디아":"cambodia",
            "몽골":"mongolia", "라오스":"laos", "괌":"guam", "사이판":"saipan",
        }
        import re, io
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            ws = wb.active
        except Exception as e:
            return jsonify({"ok": False, "error": str(e)}), 400

        headers = [str(c.value or "").strip() for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        # B열~이 국가명이므로 인덱스 매핑
        col_country = {}
        for i, h in enumerate(headers[1:], start=1):
            for kor, code in _KTO_CMAP.items():
                if kor in h:
                    col_country[i] = code
                    break

        if not col_country:
            return jsonify({"ok": False, "error": "국가 헤더를 찾을 수 없습니다. 첫 행에 '베트남', '태국' 등 한국어 국가명이 있어야 합니다."}), 400

        conn = get_db()
        saved = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            raw = str(row[0]).strip()
            # 기간 파싱: "2024년 01월" / "202401" / "2024-01" / "2024.01"
            m = re.search(r'(\d{4})[^\d]?(\d{2})', raw)
            if not m:
                continue
            ym = f"{m.group(1)}-{m.group(2)}"
            for col_i, code in col_country.items():
                if col_i >= len(row):
                    continue
                try:
                    vis = int(float(str(row[col_i]).replace(",", "")))
                except (TypeError, ValueError):
                    continue
                if vis <= 0:
                    continue
                conn.execute(
                    """INSERT INTO tourism_monthly (year_month, country, visitors, source, fetched_at)
                       VALUES (?, ?, ?, 'kto', datetime('now','localtime'))
                       ON CONFLICT(year_month, country) DO UPDATE
                       SET visitors=excluded.visitors, source=excluded.source,
                           fetched_at=excluded.fetched_at""",
                    (ym, code, vis)
                )
                saved += 1
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "saved": saved})

    # simple 모드: A=YYYY-MM, B=국가코드, C=방문자수
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 400

    conn = get_db()
    saved = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        ym  = str(row[0]).strip()
        cod = str(row[1]).strip() if row[1] else country
        try:
            vis = int(row[2])
        except (TypeError, ValueError):
            continue
        if not ym or not cod or vis <= 0:
            continue
        conn.execute(
            """INSERT INTO tourism_monthly (year_month, country, visitors, source, fetched_at)
               VALUES (?, ?, ?, 'upload', datetime('now','localtime'))
               ON CONFLICT(year_month, country) DO UPDATE
               SET visitors=excluded.visitors, source=excluded.source,
                   fetched_at=excluded.fetched_at""",
            (ym, cod, vis)
        )
        saved += 1
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "saved": saved})
